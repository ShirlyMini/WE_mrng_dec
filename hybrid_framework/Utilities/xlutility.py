# get row
# get col
# read cell
# write cell
#color
import openpyxl
from openpyxl.styles import PatternFill

def GetRow(xl_path, sheet):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    return sh.max_row

def GetCol(xl_path, sheet):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    return sh.max_column

def ReadCell(xl_path, sheet, r, c):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    return sh.cell(r,c).value

def WriteCell(xl_path, sheet, r, c, data):
    wb = openpyxl.load_workbook(xl_path)
    sh = wb[sheet]
    sh.cell(r,c).value=data
    wb.save(xl_path)

def FillColor(xl_path, sheet, r, c, color):
    if color=="green":
        green = PatternFill(start_color="34eb34", end_color="34eb34", fill_type='solid')
        wb = openpyxl.load_workbook(xl_path)
        sh = wb[sheet]
        sh.cell(r,c).fill = green
        wb.save(xl_path)
    elif color=="red":
        red = PatternFill(start_color="e30909", end_color="e30909", fill_type='solid')
        wb = openpyxl.load_workbook(xl_path)
        sh = wb[sheet]
        sh.cell(r, c).fill = red
        wb.save(xl_path)

