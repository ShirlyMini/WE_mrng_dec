# install openpyxl
import openpyxl
from openpyxl.styles import PatternFill
import os
location = fr'{os.getcwd()}\xl_test_data.xlsx'

###############create xl and writing data
# open xl/spreadsheet/workbook-->sheet--> cell(row/col)
# wb = openpyxl.Workbook()
#
# sh = wb['Sheet']
# sh.cell(1,1).value='name'
# sh.cell(2,1).value='pooja'
# sh.cell(3,1).value='nikil'
# sh.cell(4,1).value='ram'
# wb.create_sheet('Sheet123',2)
# sh = wb['Sheet123']
# for r in range(1, 6):
#     for c in range(1, 6):
#         sh.cell(r,c).value='NA'
#         red = PatternFill(start_color='e00d0d',# red
#                     end_color='e00d0d',
#                     fill_type='solid')
#         sh.cell(r, c).fill = red

# wb.save(location)####mandatory

############################reading cells

# wb_obj = openpyxl.load_workbook(location)
# sh = wb_obj['Sheet123']
# print(sh.cell(1,1).value)
# print(sh.cell(2,1).value)
# print(sh.cell(3,1).value)
# print(sh.cell(4,1).value)
#
# print(sh.max_row)
# print(sh.max_column)
#
# for r in range(1,sh.max_row+1):
#     for c in range(1, sh.max_column + 1):
#         print(sh.cell(r,c).value,end='\t')
#     print('\n')

wb = openpyxl.load_workbook(location)
sh=wb['Sheet']
count = 0
# for r in range(1, sh.max_row+1):
#     if sh.cell(r, 1).value == 'pooja':
#         count=count+1
# print('freq :',count)

# for r in range(1, sh.max_row+1):
#
#     if sh.cell(r,1).value == 'pooja':
#         count = count + 1
#         if count==3:
#             print(sh.cell(r,1).value)
#             print(sh.cell(r,2).value)
#             break


for r in range(1, sh.max_row+1):
    if sh.cell(r,1).value == 'pooja' and sh.cell(r,2).value == 'b3':
        print(f'{sh.cell(r,1).value} : {sh.cell(r,2).value}')
